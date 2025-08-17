# -*- coding: utf-8 -*-
"""
Excel Database Manager
Handles all Excel file operations for the POS system
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, timedelta
import uuid

class ExcelManager:
    def __init__(self, data_dir="data"):
        self.data_dir = data_dir
        self.ensure_data_directory()
        self.init_excel_files()
    
    def ensure_data_directory(self):
        """Create data directory if it doesn't exist"""
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
    
    def get_file_path(self, filename):
        """Get full path for Excel file"""
        return os.path.join(self.data_dir, filename)
    
    def init_excel_files(self):
        """Initialize all Excel files with proper structure"""
        self.init_products_file()
        self.init_customers_file()
        self.init_suppliers_file()
        self.init_sales_file()
        self.init_purchases_file()
        self.init_license_file()
    
    def init_products_file(self):
        """Initialize Products.xlsx file"""
        filepath = self.get_file_path("Products.xlsx")
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "المنتجات"
            
            # Headers in Arabic
            headers = ["كود المنتج", "اسم المنتج", "الباركود", "سعر التكلفة", "سعر البيع", 
                      "الكمية المتاحة", "الحد الأدنى للمخزون", "تاريخ الانتهاء", "تاريخ الإضافة"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, name="Arial Unicode MS")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(filepath)
    
    def init_customers_file(self):
        """Initialize Customers.xlsx file"""
        filepath = self.get_file_path("Customers.xlsx")
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "العملاء"
            
            headers = ["كود العميل", "اسم العميل", "رقم الهاتف", "العنوان", "مجموعة العميل",
                      "الرصيد المدين", "تاريخ التسجيل", "ملاحظات"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, name="Arial Unicode MS")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(filepath)
    
    def init_suppliers_file(self):
        """Initialize Suppliers.xlsx file"""
        filepath = self.get_file_path("Suppliers.xlsx")
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "الموردين"
            
            headers = ["كود المورد", "اسم المورد", "رقم الهاتف", "العنوان", "جدول التسليم",
                      "الرصيد الدائن", "تاريخ التسجيل", "ملاحظات"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, name="Arial Unicode MS")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(filepath)
    
    def init_sales_file(self):
        """Initialize Sales.xlsx file"""
        filepath = self.get_file_path("Sales.xlsx")
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "المبيعات"
            
            headers = ["رقم الفاتورة", "تاريخ البيع", "كود العميل", "اسم العميل", "كود المنتج",
                      "اسم المنتج", "الكمية", "سعر الوحدة", "الخصم", "الإجمالي", "المبلغ المدفوع",
                      "المتبقي", "طريقة الدفع", "ملاحظات"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, name="Arial Unicode MS")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(filepath)
    
    def init_purchases_file(self):
        """Initialize Purchases.xlsx file"""
        filepath = self.get_file_path("Purchases.xlsx")
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "المشتريات"
            
            headers = ["رقم الفاتورة", "تاريخ الشراء", "كود المورد", "اسم المورد", "كود المنتج",
                      "اسم المنتج", "الكمية", "سعر الوحدة", "الإجمالي", "المبلغ المدفوع",
                      "المتبقي", "تاريخ الاستحقاق", "ملاحظات"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, name="Arial Unicode MS")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(filepath)
    
    def init_license_file(self):
        """Initialize License.xlsx file"""
        filepath = self.get_file_path("License.xlsx")
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "الترخيص"
            
            headers = ["معرف الجهاز", "تاريخ البداية", "تاريخ الانتهاء", "الأيام المتبقية", "حالة الترخيص"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, name="Arial Unicode MS")
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(filepath)
    
    # Product Management Methods
    def add_product(self, product_data):
        """Add new product to Products.xlsx"""
        filepath = self.get_file_path("Products.xlsx")
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Generate product code if not provided
        if not product_data.get('كود المنتج'):
            product_data['كود المنتج'] = f"P{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Add current date
        product_data['تاريخ الإضافة'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add row
        row_data = [
            product_data.get('كود المنتج', ''),
            product_data.get('اسم المنتج', ''),
            product_data.get('الباركود', ''),
            product_data.get('سعر التكلفة', 0),
            product_data.get('سعر البيع', 0),
            product_data.get('الكمية المتاحة', 0),
            product_data.get('الحد الأدنى للمخزون', 0),
            product_data.get('تاريخ الانتهاء', ''),
            product_data.get('تاريخ الإضافة', '')
        ]
        
        ws.append(row_data)
        wb.save(filepath)
        return product_data['كود المنتج']
    
    def get_products(self):
        """Get all products from Products.xlsx"""
        filepath = self.get_file_path("Products.xlsx")
        if not os.path.exists(filepath):
            return []
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        products = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                product = dict(zip(headers, row))
                products.append(product)
        
        return products
    
    def get_low_stock_products(self):
        """Get products with low stock"""
        products = self.get_products()
        low_stock = []
        
        for product in products:
            current_qty = product.get('الكمية المتاحة', 0) or 0
            min_qty = product.get('الحد الأدنى للمخزون', 0) or 0
            
            if current_qty <= min_qty:
                low_stock.append(product)
        
        return low_stock
    
    def get_expiring_products(self, days_ahead=30):
        """Get products expiring within specified days"""
        products = self.get_products()
        expiring = []
        
        cutoff_date = datetime.now() + timedelta(days=days_ahead)
        
        for product in products:
            expiry_date = product.get('تاريخ الانتهاء', '')
            if expiry_date:
                try:
                    expiry = datetime.strptime(expiry_date, '%Y-%m-%d')
                    if expiry <= cutoff_date:
                        expiring.append(product)
                except ValueError:
                    continue
        
        return expiring
    
    # Customer Management Methods
    def add_customer(self, customer_data):
        """Add new customer to Customers.xlsx"""
        filepath = self.get_file_path("Customers.xlsx")
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Generate customer code if not provided
        if not customer_data.get('كود العميل'):
            customer_data['كود العميل'] = f"C{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        customer_data['تاريخ التسجيل'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        row_data = [
            customer_data.get('كود العميل', ''),
            customer_data.get('اسم العميل', ''),
            customer_data.get('رقم الهاتف', ''),
            customer_data.get('العنوان', ''),
            customer_data.get('مجموعة العميل', 'عادي'),
            customer_data.get('الرصيد المدين', 0),
            customer_data.get('تاريخ التسجيل', ''),
            customer_data.get('ملاحظات', '')
        ]
        
        ws.append(row_data)
        wb.save(filepath)
        return customer_data['كود العميل']
    
    def get_customers(self):
        """Get all customers from Customers.xlsx"""
        filepath = self.get_file_path("Customers.xlsx")
        if not os.path.exists(filepath):
            return []
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        customers = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                customer = dict(zip(headers, row))
                customers.append(customer)
        
        return customers
    
    # Sales Management Methods
    def add_sale(self, sale_data):
        """Add new sale to Sales.xlsx"""
        filepath = self.get_file_path("Sales.xlsx")
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Generate invoice number if not provided
        if not sale_data.get('رقم الفاتورة'):
            sale_data['رقم الفاتورة'] = f"INV{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        sale_data['تاريخ البيع'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        row_data = [
            sale_data.get('رقم الفاتورة', ''),
            sale_data.get('تاريخ البيع', ''),
            sale_data.get('كود العميل', ''),
            sale_data.get('اسم العميل', ''),
            sale_data.get('كود المنتج', ''),
            sale_data.get('اسم المنتج', ''),
            sale_data.get('الكمية', 0),
            sale_data.get('سعر الوحدة', 0),
            sale_data.get('الخصم', 0),
            sale_data.get('الإجمالي', 0),
            sale_data.get('المبلغ المدفوع', 0),
            sale_data.get('المتبقي', 0),
            sale_data.get('طريقة الدفع', 'نقدي'),
            sale_data.get('ملاحظات', '')
        ]
        
        ws.append(row_data)
        wb.save(filepath)
        
        # Update product stock
        self.update_product_stock(sale_data.get('كود المنتج'), -sale_data.get('الكمية', 0))
        
        return sale_data['رقم الفاتورة']
    
    def update_product_stock(self, product_code, quantity_change):
        """Update product stock quantity"""
        filepath = self.get_file_path("Products.xlsx")
        wb = load_workbook(filepath)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == product_code:
                current_qty = ws.cell(row=row, column=6).value or 0
                new_qty = current_qty + quantity_change
                ws.cell(row=row, column=6, value=max(0, new_qty))
                break
        
        wb.save(filepath)
    
    # License Management Methods
    def get_license_info(self):
        """Get license information"""
        filepath = self.get_file_path("License.xlsx")
        if not os.path.exists(filepath):
            return None
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        if ws.max_row > 1:
            row = ws[2]
            return {
                'device_id': row[0].value,
                'start_date': row[1].value,
                'expiry_date': row[2].value,
                'days_left': row[3].value,
                'status': row[4].value
            }
        return None
    
    def set_license(self, device_id, days):
        """Set license for specified days"""
        filepath = self.get_file_path("License.xlsx")
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Clear existing data
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        
        start_date = datetime.now()
        expiry_date = start_date + timedelta(days=days)
        
        row_data = [
            device_id,
            start_date.strftime('%Y-%m-%d'),
            expiry_date.strftime('%Y-%m-%d'),
            days,
            'نشط'
        ]
        
        ws.append(row_data)
        wb.save(filepath)
    
    # Search Methods
    def search_products(self, search_term):
        """Search products by name or code"""
        products = self.get_products()
        results = []
        
        search_term = search_term.lower()
        for product in products:
            name = str(product.get('اسم المنتج', '')).lower()
            code = str(product.get('كود المنتج', '')).lower()
            barcode = str(product.get('الباركود', '')).lower()
            
            if (search_term in name or search_term in code or search_term in barcode):
                results.append(product)
        
        return results
    
    def search_customers(self, search_term):
        """Search customers by name or phone"""
        customers = self.get_customers()
        results = []
        
        search_term = search_term.lower()
        for customer in customers:
            name = str(customer.get('اسم العميل', '')).lower()
            phone = str(customer.get('رقم الهاتف', '')).lower()
            
            if search_term in name or search_term in phone:
                results.append(customer)
        
        return results
    
    # Reports Methods
    def get_daily_sales_report(self, date=None):
        """Get daily sales report"""
        if date is None:
            date = datetime.now().strftime('%Y-%m-%d')
        
        filepath = self.get_file_path("Sales.xlsx")
        if not os.path.exists(filepath):
            return []
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        sales = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:  # Check if date exists
                sale_date = str(row[1])[:10]  # Get date part only
                if sale_date == date:
                    sale = dict(zip(headers, row))
                    sales.append(sale)
        
        return sales
    
    def get_monthly_sales_report(self, year_month=None):
        """Get monthly sales report"""
        if year_month is None:
            year_month = datetime.now().strftime('%Y-%m')
        
        filepath = self.get_file_path("Sales.xlsx")
        if not os.path.exists(filepath):
            return []
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        sales = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:  # Check if date exists
                sale_date = str(row[1])[:7]  # Get year-month part only
                if sale_date == year_month:
                    sale = dict(zip(headers, row))
                    sales.append(sale)
        
        return sales